"""
OPENBOX - TALLER 4 CON 1000 CLIENTES RANDOM
Independiente de tu simulación original.
"""

import random
import openpyxl
from openpyxl.styles import Font, Alignment
import os
import variables


# =============================
# GENERAR 1000 CLIENTES RANDOM
# =============================



print("Usando este archivo variables.py:")
print(os.path.abspath(variables.__file__))

def generar_clientes(n=1000):
    clientes = []

    for _ in range(n):

        # Distribución de artículos (como en supermercados reales)
        r = random.random()
        if r < 0.60:         # 60% lleva pocos artículos
            articulos = random.randint(1, 15)
        elif r < 0.90:       # 30% intermedio
            articulos = random.randint(16, 40)
        else:                # 10% compras grandes
            articulos = random.randint(41, 80)

        # Tiempo de cobro
        t_cobro = random.randint(15, 30)

        clientes.append((articulos, t_cobro))

    return clientes


# =============================
# CÁLCULO DE UNA RÉPLICA
# =============================
def ejecutar_replicas(s):
    resultados = []

    REP = variables.REPLICAS_OPENBOX
    OBJ = variables.OBJETIVO_SLA_SEG

    for replica in range(1, REP + 1):

        tiempos = []
        clientes = generar_clientes(1000)

        # Para el cálculo de λ y μ
        total_clientes = len(clientes)

        # Simular s cajas
        for _ in range(s):

            # Cajero aleatorio
            cajero = random.choice(["Experto", "Principiante"])

            if cajero == "Experto":
                t_escaneo = random.uniform(2.5, 4.0)
            else:
                t_escaneo = random.uniform(4.5, 7.0)

            # Servicio del cliente
            for articulos, cobro in clientes:
                t_serv = articulos * t_escaneo + cobro
                tiempos.append(t_serv)

        # ===========================
        # MÉTRICAS DEL TALLER
        # ===========================
        total = sum(tiempos)
        T_prom = total / len(tiempos)       # tiempo por cliente

        λ = total_clientes / (total / 60)   # clientes/min
        μ = 1 / (T_prom / 60)               # clientes/min

        ρ = λ / (s * μ)

        W = T_prom / 60
        Wq = max(0, W - 1 / μ)

        L = λ * W
        Lq = λ * Wq

        SLA = sum(1 for t in tiempos if t <= OBJ)
        SLA = (SLA / len(tiempos)) * 100

        # Costo total
        CT = (
            variables.COSTO_CAJA * s * W +
            variables.COSTO_ESPERA * (total / 60) +
            variables.COSTO_SLA * (100 - SLA)
        )

        resultados.append({
            "s": s,
            "replica": replica,
            "clientes": len(tiempos),
            "T_prom": T_prom,
            "W": W,
            "Wq": Wq,
            "L": L,
            "Lq": Lq,
            "rho": ρ,
            "SLA": SLA,
            "CT": CT
        })

    return resultados


# =============================
# EXPORTAR A EXCEL
# =============================
def exportar_excel(matriz):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Taller4"

    headers = ["s", "replica", "clientes", "T_prom", "W", "Wq", "L", "Lq", "rho", "%SLA", "CT"]

    for col, h in enumerate(headers, start=1):
        c = ws.cell(1, col, h)
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal="center")

    for i, fila in enumerate(matriz, start=2):
        ws.cell(i, 1, fila["s"])
        ws.cell(i, 2, fila["replica"])
        ws.cell(i, 3, fila["clientes"])
        ws.cell(i, 4, round(fila["T_prom"], 2))
        ws.cell(i, 5, round(fila["W"], 3))
        ws.cell(i, 6, round(fila["Wq"], 3))
        ws.cell(i, 7, round(fila["L"], 3))
        ws.cell(i, 8, round(fila["Lq"], 3))
        ws.cell(i, 9, round(fila["rho"], 3))
        ws.cell(i,10, round(fila["SLA"], 2))
        ws.cell(i,11, round(fila["CT"], 2))

    wb.save("openbox.xlsx")
    print("\n✔ Archivo generado: openbox.xlsx")




def ejecutar_openbox():
    matriz = []
    print("=== OPENBOX 1000 CLIENTES RANDOM ===\n")

    for s in variables.VALORES_S:
        print(f"Simulando s = {s} cajas…")
        filas = ejecutar_replicas(s)
        matriz.extend(filas)

    exportar_excel(matriz)
    print("\n✔ Simulación del Taller 4 completada.")


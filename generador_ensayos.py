"""
Generador de Ensayos para Simulación de Cajas
Genera múltiples ensayos sin transcurrir tiempo real, solo cálculos matemáticos
"""
import random
from datetime import datetime
import variables
import variables

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    TIENE_OPENPYXL = True
except ImportError:
    TIENE_OPENPYXL = False
    print("⚠️ Advertencia: openpyxl no está instalado. Instalando...")
    import subprocess
    import sys
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    TIENE_OPENPYXL = True


class EnsayoSimulador:
    """Simulador que calcula tiempos sin esperar (sin time.sleep)"""
    
    def __init__(self, caso_prueba="Uniforme"):
        self.caso_prueba = caso_prueba
        self.config = variables.CASOS_PRUEBA[caso_prueba]
        self.resultados = []
    
    def generar_cliente_objetivo(self):
        """Genera un cliente objetivo con artículos aleatorios"""
        config_obj = self.config['cliente_objetivo']
        articulos = random.randint(config_obj['articulos_min'], config_obj['articulos_max'])
        tiempo_cobro = random.randint(self.config['tiempo_cobro_min'], self.config['tiempo_cobro_max'])
        return articulos, tiempo_cobro
    
    def generar_personas_caja(self, caja_config):
        """Genera personas para una caja según su configuración"""
        cantidad = random.randint(caja_config['personas_min'], caja_config['personas_max'])
        personas = []
        for _ in range(cantidad):
            r = random.random()
            articulos = None
            for prob, min_art, max_art in caja_config['articulos_distribucion']:
                if r <= prob:
                    articulos = random.randint(min_art, max_art)
                    break
            # Si no se asignó (por algún error), usar valores por defecto
            if articulos is None:
                articulos = random.randint(1, 10)
            tiempo_cobro = random.randint(self.config['tiempo_cobro_min'], self.config['tiempo_cobro_max'])
            personas.append((articulos, tiempo_cobro))
        
        return personas
    
    def calcular_tiempo_hasta_objetivo(self, personas, cliente_objetivo, tiempo_escaneo):
        """Calcula el tiempo total hasta atender al cliente objetivo"""
        articulos_objetivo, tiempo_cobro_objetivo = cliente_objetivo
        
        # Agregar el cliente objetivo al final
        todas_personas = personas + [(articulos_objetivo, tiempo_cobro_objetivo)]
        
        tiempo_total = 0
        for articulos, tiempo_cobro in todas_personas:
            tiempo_persona = articulos * tiempo_escaneo + tiempo_cobro
            tiempo_total += tiempo_persona
        
        return tiempo_total
    
    def ejecutar_ensayo_simple(self, numero_ensayo):
        """Ejecuta un ensayo completo y retorna el resultado"""
        
        # Generar cliente objetivo (mismo para las 3 cajas)
        cliente_objetivo = self.generar_cliente_objetivo()
        articulos_objetivo, tiempo_cobro_objetivo = cliente_objetivo
        
        # Configuración de cajas desde el caso de prueba
        cajas = []
        for i, caja_config in enumerate(self.config['cajas']):
            # Asignar cajero aleatoriamente
            cajero_aleatorio = random.choice(['Experto', 'Principiante'])
            
            # Determinar tiempos de escaneo según el cajero
            if cajero_aleatorio == 'Experto':
                tiempo_escaneo = random.uniform(2.5, 4.0)
            else:  # Principiante
                tiempo_escaneo = random.uniform(4.5, 7.0)
            
            caja = {
                'id': i+1,
                'tipo': caja_config['tipo'],
                'cajero': cajero_aleatorio,
                'tiempo_escaneo': tiempo_escaneo,
                'express': caja_config['express'],
                'config': caja_config
            }
            cajas.append(caja)
        
        # Calcular tiempo para cada caja
        tiempos = []
        for caja in cajas:
            personas = self.generar_personas_caja(caja['config'])
            
            # Calcular cantidad total de productos en esta caja
            total_productos = sum(persona[0] for persona in personas)  # persona[0] es la cantidad de artículos
            
            # Calcular promedio del tiempo de cobro para esta caja
            tiempos_cobro_caja = [persona[1] for persona in personas]  # persona[1] es el tiempo de cobro
            promedio_cobro_caja = sum(tiempos_cobro_caja) / len(tiempos_cobro_caja) if tiempos_cobro_caja else 0
            
            tiempo = self.calcular_tiempo_hasta_objetivo(
                personas, 
                cliente_objetivo, 
                caja['tiempo_escaneo']
            )
            tiempos.append({
                'caja_id': caja['id'],
                'tipo_caja': caja['tipo'],
                'cajero': caja['cajero'],
                'tiempo_escaneo': round(caja['tiempo_escaneo'], 2),
                'num_personas': len(personas),
                'total_productos': total_productos,
                'promedio_cobro': round(promedio_cobro_caja, 2),
                'tiempo_total': round(tiempo, 2)
            })
        
        # Encontrar la caja ganadora (menor tiempo)
        tiempos_ordenados = sorted(tiempos, key=lambda x: x['tiempo_total'])
        ganadora = tiempos_ordenados[0]
        
        # Clasificar ganador simplificado
        if ganadora['tipo_caja'] == 'Express':
            tipo_ganador = 'Express'
        else:
            tipo_ganador = 'Normal'
        
        return {
            'ensayo': numero_ensayo,
            'articulos_objetivo': articulos_objetivo,
            'tiempo_cobro_objetivo': tiempo_cobro_objetivo,
            'caja_ganadora_id': ganadora['caja_id'],
            'tipo_ganador': tipo_ganador,
            'tiempo_ganador': ganadora['tiempo_total'],
            'caja1_tiempo': tiempos[0]['tiempo_total'],
            'caja1_personas': tiempos[0]['num_personas'],
            'caja1_productos': tiempos[0]['total_productos'],
            'caja1_tiempo_escaneo': tiempos[0]['tiempo_escaneo'],
            'caja1_promedio_cobro': tiempos[0]['promedio_cobro'],
            'caja2_tiempo': tiempos[1]['tiempo_total'],
            'caja2_personas': tiempos[1]['num_personas'],
            'caja2_productos': tiempos[1]['total_productos'],
            'caja2_tiempo_escaneo': tiempos[1]['tiempo_escaneo'],
            'caja2_promedio_cobro': tiempos[1]['promedio_cobro'],
            'caja3_tiempo': tiempos[2]['tiempo_total'],
            'caja3_personas': tiempos[2]['num_personas'],
            'caja3_productos': tiempos[2]['total_productos'],
            'caja3_tiempo_escaneo': tiempos[2]['tiempo_escaneo'],
            'caja3_promedio_cobro': tiempos[2]['promedio_cobro']
        }
    
    def generar_ensayos(self, cantidad=1000):
        """Genera múltiples ensayos y guarda en CSV"""
        print(f"Generando {cantidad} ensayos instantáneos...")
        
        resultados = []
        for i in range(1, cantidad + 1):
            resultado = self.ejecutar_ensayo_simple(i)
            resultados.append(resultado)
            
            if i % 100 == 0:
                print(f"  Progreso: {i}/{cantidad} ensayos completados")
        
        self.resultados = resultados
        return resultados
    
    def exportar_excel(self, nombre_archivo=None):
        """Exporta los resultados a un archivo Excel (.xlsx)"""
        if not self.resultados:
            print("No hay resultados para exportar")
            return
        
        if nombre_archivo is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            nombre_archivo = f"ensayos_cajas_{timestamp}.xlsx"
        
        # Asegurar extensión .xlsx
        if not nombre_archivo.endswith('.xlsx'):
            nombre_archivo += '.xlsx'
        
        # Usar path absoluto para asegurar que se guarde en el directorio correcto
        import os
        nombre_archivo = os.path.join(os.getcwd(), nombre_archivo)
        
        # Crear libro de Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Ensayos"
        
        # Calcular promedio del tiempo de cobro objetivo
        tiempos_cobro = [resultado['tiempo_cobro_objetivo'] for resultado in self.resultados]
        promedio_cobro = sum(tiempos_cobro) / len(tiempos_cobro) if tiempos_cobro else 0
        
        # Definir encabezados
        encabezados = [
            'Ensayo',
            'Articulos_Objetivo',
            'Tiempo_Cobro_Objetivo',
            'Promedio_Cobro_Global',
            'Caja_Ganadora_ID',
            'Tipo_Ganador',
            'Tiempo_Ganador',
            'Caja1_Tiempo',
            'Caja1_Personas',
            'Caja1_Productos',
            'Caja1_Tiempo_Escaneo',
            'Caja1_Promedio_Cobro',
            'Caja2_Tiempo',
            'Caja2_Personas',
            'Caja2_Productos',
            'Caja2_Tiempo_Escaneo',
            'Caja2_Promedio_Cobro',
            'Caja3_Tiempo',
            'Caja3_Personas',
            'Caja3_Productos',
            'Caja3_Tiempo_Escaneo',
            'Caja3_Promedio_Cobro'
        ]
        
        # Estilo para encabezados
        header_fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        center_alignment = Alignment(horizontal="center", vertical="center")
        
        # Escribir encabezados
        for col_idx, encabezado in enumerate(encabezados, start=1):
            cell = ws.cell(row=1, column=col_idx, value=encabezado)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_alignment
        
        # Escribir datos
        for row_idx, resultado in enumerate(self.resultados, start=2):
            ws.cell(row=row_idx, column=1, value=resultado['ensayo'])
            ws.cell(row=row_idx, column=2, value=resultado['articulos_objetivo'])
            ws.cell(row=row_idx, column=3, value=resultado['tiempo_cobro_objetivo'])
            ws.cell(row=row_idx, column=4, value=round(promedio_cobro, 2))  # Promedio global
            ws.cell(row=row_idx, column=5, value=resultado['caja_ganadora_id'])
            ws.cell(row=row_idx, column=6, value=resultado['tipo_ganador'])
            ws.cell(row=row_idx, column=7, value=round(resultado['tiempo_ganador'], 3))
            ws.cell(row=row_idx, column=8, value=round(resultado['caja1_tiempo'], 3))
            ws.cell(row=row_idx, column=9, value=resultado['caja1_personas'])
            ws.cell(row=row_idx, column=10, value=resultado['caja1_productos'])
            ws.cell(row=row_idx, column=11, value=round(resultado['caja1_tiempo_escaneo'], 2))
            ws.cell(row=row_idx, column=12, value=round(resultado['caja1_promedio_cobro'], 2))
            ws.cell(row=row_idx, column=13, value=round(resultado['caja2_tiempo'], 3))
            ws.cell(row=row_idx, column=14, value=resultado['caja2_personas'])
            ws.cell(row=row_idx, column=15, value=resultado['caja2_productos'])
            ws.cell(row=row_idx, column=16, value=round(resultado['caja2_tiempo_escaneo'], 2))
            ws.cell(row=row_idx, column=17, value=round(resultado['caja2_promedio_cobro'], 2))
            ws.cell(row=row_idx, column=18, value=round(resultado['caja3_tiempo'], 3))
            ws.cell(row=row_idx, column=19, value=resultado['caja3_personas'])
            ws.cell(row=row_idx, column=20, value=resultado['caja3_productos'])
            ws.cell(row=row_idx, column=21, value=round(resultado['caja3_tiempo_escaneo'], 2))
            ws.cell(row=row_idx, column=22, value=round(resultado['caja3_promedio_cobro'], 2))
            
            # Resaltar la caja ganadora (columna de tiempo ganadora)
            ganadora_col = 5 + (resultado['caja_ganadora_id'] - 1) * 5 + 3  # Ajustado para nuevas columnas
            cell_ganadora = ws.cell(row=row_idx, column=ganadora_col)
            cell_ganadora.fill = PatternFill(start_color="FFEB3B", end_color="FFEB3B", fill_type="solid")
        
        # Ajustar ancho de columnas
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width
        
        # Crear hoja de estadísticas
        ws_stats = wb.create_sheet("Estadísticas")
        ws_stats.title = "Estadísticas"
        
        # Calcular estadísticas
        total = len(self.resultados)
        ganadores_express = sum(1 for r in self.resultados if r['tipo_ganador'] == 'Express')
        ganadores_normal = total - ganadores_express
        ganadores_caja1 = sum(1 for r in self.resultados if r['caja_ganadora_id'] == 1)
        ganadores_caja2 = sum(1 for r in self.resultados if r['caja_ganadora_id'] == 2)
        ganadores_caja3 = sum(1 for r in self.resultados if r['caja_ganadora_id'] == 3)
        
        tiempo_promedio_caja1 = sum(r['caja1_tiempo'] for r in self.resultados) / total
        tiempo_promedio_caja2 = sum(r['caja2_tiempo'] for r in self.resultados) / total
        tiempo_promedio_caja3 = sum(r['caja3_tiempo'] for r in self.resultados) / total
        
        # Calcular promedios de productos por caja
        productos_promedio_caja1 = sum(r['caja1_productos'] for r in self.resultados) / total
        productos_promedio_caja2 = sum(r['caja2_productos'] for r in self.resultados) / total
        productos_promedio_caja3 = sum(r['caja3_productos'] for r in self.resultados) / total
        
        # Escribir estadísticas
        ws_stats.cell(row=1, column=1, value="ESTADÍSTICAS RESUMIDAS").font = Font(bold=True, size=14)
        ws_stats.cell(row=2, column=1, value=f"Total de ensayos: {total}")
        
        ws_stats.cell(row=4, column=1, value="GANADORES POR TIPO:").font = Font(bold=True)
        ws_stats.cell(row=5, column=1, value=f"Cajas Normales (1+2): {ganadores_normal} ({ganadores_normal/total*100:.1f}%)")
        ws_stats.cell(row=6, column=1, value=f"Caja Express (3): {ganadores_express} ({ganadores_express/total*100:.1f}%)")
        
        ws_stats.cell(row=8, column=1, value="GANADORES POR CAJA ESPECÍFICA:").font = Font(bold=True)
        ws_stats.cell(row=9, column=1, value=f"Caja 1 (Normal): {ganadores_caja1} ({ganadores_caja1/total*100:.1f}%)")
        ws_stats.cell(row=10, column=1, value=f"Caja 2 (Normal): {ganadores_caja2} ({ganadores_caja2/total*100:.1f}%)")
        ws_stats.cell(row=11, column=1, value=f"Caja 3 (Express): {ganadores_caja3} ({ganadores_caja3/total*100:.1f}%)")
        
        ws_stats.cell(row=13, column=1, value="TIEMPOS PROMEDIO:").font = Font(bold=True)
        ws_stats.cell(row=14, column=1, value=f"Caja 1: {tiempo_promedio_caja1:.1f} segundos")
        ws_stats.cell(row=15, column=1, value=f"Caja 2: {tiempo_promedio_caja2:.1f} segundos")
        ws_stats.cell(row=16, column=1, value=f"Caja 3: {tiempo_promedio_caja3:.1f} segundos")
        
        ws_stats.cell(row=18, column=1, value="PRODUCTOS PROMEDIO POR CAJA:").font = Font(bold=True)
        ws_stats.cell(row=19, column=1, value=f"Caja 1: {productos_promedio_caja1:.1f} productos")
        ws_stats.cell(row=20, column=1, value=f"Caja 2: {productos_promedio_caja2:.1f} productos")
        ws_stats.cell(row=21, column=1, value=f"Caja 3: {productos_promedio_caja3:.1f} productos")
        
        # Ajustar ancho de columna
        ws_stats.column_dimensions['A'].width = 50
        
        # Guardar el archivo Excel
        wb.save(nombre_archivo)
        
        print(f"\n✓ Archivo Excel exportado: {nombre_archivo}")
        return nombre_archivo
    
    def mostrar_estadisticas(self):
        """Muestra estadísticas de los ensayos"""
        if not self.resultados:
            print("No hay resultados para analizar")
            return
        
        total = len(self.resultados)
        
        # Contar ganadores por tipo
        ganadores_express = sum(1 for r in self.resultados if r['tipo_ganador'] == 'Express')
        ganadores_normal = sum(1 for r in self.resultados if r['tipo_ganador'] == 'Normal')
        
        # Contar por caja específica
        ganadores_caja1 = sum(1 for r in self.resultados if r['caja_ganadora_id'] == 1)
        ganadores_caja2 = sum(1 for r in self.resultados if r['caja_ganadora_id'] == 2)
        ganadores_caja3 = sum(1 for r in self.resultados if r['caja_ganadora_id'] == 3)
        
        # Calcular tiempos promedio por caja
        tiempo_promedio_caja1 = sum(r['caja1_tiempo'] for r in self.resultados) / total
        tiempo_promedio_caja2 = sum(r['caja2_tiempo'] for r in self.resultados) / total
        tiempo_promedio_caja3 = sum(r['caja3_tiempo'] for r in self.resultados) / total
        
        print("\n" + "="*60)
        print("ESTADÍSTICAS DE LOS ENSAYOS")
        print("="*60)
        print(f"Total de ensayos: {total}")
        print("\nGanadores por TIPO de caja:")
        print(f"  Cajas Normales (1+2):  {ganadores_normal} ({ganadores_normal/total*100:.1f}%)")
        print(f"  Caja Express (3):      {ganadores_express} ({ganadores_express/total*100:.1f}%)")
        print("\nGanadores por caja específica:")
        print(f"  Caja 1 (Normal):        {ganadores_caja1} ({ganadores_caja1/total*100:.1f}%)")
        print(f"  Caja 2 (Normal):        {ganadores_caja2} ({ganadores_caja2/total*100:.1f}%)")
        print(f"  Caja 3 (Express):       {ganadores_caja3} ({ganadores_caja3/total*100:.1f}%)")
        print("\nTiempos promedio por caja:")
        print(f"  Caja 1: {tiempo_promedio_caja1:.1f} segundos")
        print(f"  Caja 2: {tiempo_promedio_caja2:.1f} segundos")
        print(f"  Caja 3: {tiempo_promedio_caja3:.1f} segundos")
        print("="*60)


def main():
    """Función principal para ejecutar los ensayos"""
    print("="*60)
    print("GENERADOR DE ENSAYOS - SIMULACIÓN DE CAJAS")
    print("="*60)
    
    simulador = EnsayoSimulador()
    
    # Generar ensayos
    simulador.generar_ensayos(variables.CASOS_PRUEBA[simulador.caso_prueba]['cantidad_ensayos'])
    
    # Mostrar estadísticas
    simulador.mostrar_estadisticas()
    
    # Exportar a Excel
    archivo = simulador.exportar_excel()
    
    print(f"\n✓ Proceso completado exitosamente!")
    print(f"  Archivo Excel generado: {archivo}")


if __name__ == "__main__":
    main()